#!/usr/bin/env python3
"""
DNA to Amino Acid Translator
Translates DNA sequences in all 6 reading frames and identifies the best one.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os

# Standard codon table
CODON_TABLE = {
    'TTT': 'F', 'TTC': 'F', 'TTA': 'L', 'TTG': 'L',
    'TCT': 'S', 'TCC': 'S', 'TCA': 'S', 'TCG': 'S',
    'TAT': 'Y', 'TAC': 'Y', 'TAA': '*', 'TAG': '*',
    'TGT': 'C', 'TGC': 'C', 'TGA': '*', 'TGG': 'W',
    'CTT': 'L', 'CTC': 'L', 'CTA': 'L', 'CTG': 'L',
    'CCT': 'P', 'CCC': 'P', 'CCA': 'P', 'CCG': 'P',
    'CAT': 'H', 'CAC': 'H', 'CAA': 'Q', 'CAG': 'Q',
    'CGT': 'R', 'CGC': 'R', 'CGA': 'R', 'CGG': 'R',
    'ATT': 'I', 'ATC': 'I', 'ATA': 'I', 'ATG': 'M',
    'ACT': 'T', 'ACC': 'T', 'ACA': 'T', 'ACG': 'T',
    'AAT': 'N', 'AAC': 'N', 'AAA': 'K', 'AAG': 'K',
    'AGT': 'S', 'AGC': 'S', 'AGA': 'R', 'AGG': 'R',
    'GTT': 'V', 'GTC': 'V', 'GTA': 'V', 'GTG': 'V',
    'GCT': 'A', 'GCC': 'A', 'GCA': 'A', 'GCG': 'A',
    'GAT': 'D', 'GAC': 'D', 'GAA': 'E', 'GAG': 'E',
    'GGT': 'G', 'GGC': 'G', 'GGA': 'G', 'GGG': 'G',
}

def clean_sequence(seq):
    """Remove whitespace and convert to uppercase."""
    if pd.isna(seq):
        return ""
    return ''.join(str(seq).upper().split())

def reverse_complement(seq):
    """Return the reverse complement of a DNA sequence."""
    complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G',
                  'N': 'N', 'R': 'Y', 'Y': 'R', 'S': 'S',
                  'W': 'W', 'K': 'M', 'M': 'K'}
    return ''.join(complement.get(base, 'N') for base in reversed(seq))

def translate(seq, frame=0):
    """
    Translate a DNA sequence to amino acids.
    frame: 0, 1, or 2 for the three reading frames
    Returns: (protein_sequence, stop_count, longest_orf_length)
    """
    protein = []
    seq = seq[frame:]  # Adjust for reading frame
    
    for i in range(0, len(seq) - 2, 3):
        codon = seq[i:i+3]
        if len(codon) == 3:
            aa = CODON_TABLE.get(codon, 'X')  # X for unknown
            protein.append(aa)
    
    protein_str = ''.join(protein)
    stop_count = protein_str.count('*')
    
    # Find longest ORF (sequence between stops, or from start to first stop)
    orfs = protein_str.replace('*', ' ').split()
    longest_orf = max(len(orf) for orf in orfs) if orfs else 0
    
    return protein_str, stop_count, longest_orf

def analyze_sequence(name, seq):
    """
    Analyze a DNA sequence in all 6 reading frames.
    Returns a list of results for each frame.
    """
    seq = clean_sequence(seq)
    if not seq:
        return None
    
    rev_comp = reverse_complement(seq)
    results = []
    
    # Forward frames (+1, +2, +3)
    for frame in range(3):
        protein, stops, longest_orf = translate(seq, frame)
        results.append({
            'Sequence_Name': name,
            'Frame': f'+{frame + 1}',
            'Strand': 'Forward',
            'Protein': protein,
            'Length_AA': len(protein),
            'Stop_Codons': stops,
            'Longest_ORF': longest_orf,
            'DNA_Length': len(seq)
        })
    
    # Reverse complement frames (-1, -2, -3)
    for frame in range(3):
        protein, stops, longest_orf = translate(rev_comp, frame)
        results.append({
            'Sequence_Name': name,
            'Frame': f'-{frame + 1}',
            'Strand': 'Reverse',
            'Protein': protein,
            'Length_AA': len(protein),
            'Stop_Codons': stops,
            'Longest_ORF': longest_orf,
            'DNA_Length': len(seq)
        })
    
    return results

def find_best_frame(results):
    """
    Determine the best reading frame based on:
    1. Longest ORF (primary criterion)
    2. Fewest stop codons (secondary criterion)
    """
    if not results:
        return None
    
    # Sort by longest ORF (descending), then by stop codons (ascending)
    sorted_results = sorted(results, key=lambda x: (-x['Longest_ORF'], x['Stop_Codons']))
    return sorted_results[0]['Frame']

def read_sequences(input_file):
    """
    Read sequences from Excel or CSV file.
    Expects columns: 'Name' (or 'ID' or first column) and 'Sequence' (or second column)
    """
    ext = os.path.splitext(input_file)[1].lower()
    
    if ext in ['.xlsx', '.xls']:
        df = pd.read_excel(input_file)
    elif ext == '.csv':
        df = pd.read_csv(input_file)
    elif ext == '.tsv':
        df = pd.read_csv(input_file, sep='\t')
    else:
        raise ValueError(f"Unsupported file format: {ext}")
    
    # Try to identify name and sequence columns
    df.columns = [str(c).strip() for c in df.columns]
    
    name_col = None
    seq_col = None
    
    # Look for common column names
    for col in df.columns:
        col_lower = col.lower()
        if col_lower in ['name', 'id', 'sequence_name', 'seq_name', 'gene', 'gene_name']:
            name_col = col
        elif col_lower in ['sequence', 'seq', 'dna', 'dna_sequence', 'nucleotide']:
            seq_col = col
    
    # Default to first two columns if not found
    if name_col is None:
        name_col = df.columns[0]
    if seq_col is None:
        seq_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    
    return df[[name_col, seq_col]].rename(columns={name_col: 'Name', seq_col: 'Sequence'})

def create_output(all_results, best_frames, output_file):
    """Create formatted Excel output with results."""
    wb = Workbook()
    
    # Sheet 1: Summary (best frame for each sequence)
    ws_summary = wb.active
    ws_summary.title = "Best_Frames"
    
    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary headers
    summary_headers = ['Sequence_Name', 'Best_Frame', 'Protein', 'Length_AA', 
                       'Longest_ORF', 'Stop_Codons', 'DNA_Length']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Add best frame data
    row = 2
    for seq_name, frame in best_frames.items():
        # Find the result for this sequence and frame
        for result in all_results:
            if result['Sequence_Name'] == seq_name and result['Frame'] == frame:
                ws_summary.cell(row=row, column=1, value=result['Sequence_Name'])
                ws_summary.cell(row=row, column=2, value=result['Frame'])
                ws_summary.cell(row=row, column=3, value=result['Protein'])
                ws_summary.cell(row=row, column=4, value=result['Length_AA'])
                ws_summary.cell(row=row, column=5, value=result['Longest_ORF'])
                ws_summary.cell(row=row, column=6, value=result['Stop_Codons'])
                ws_summary.cell(row=row, column=7, value=result['DNA_Length'])
                row += 1
                break
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 50
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 12
    ws_summary.column_dimensions['G'].width = 12
    
    # Sheet 2: All frames detail
    ws_detail = wb.create_sheet("All_Frames")
    
    detail_headers = ['Sequence_Name', 'Frame', 'Strand', 'Protein', 'Length_AA', 
                      'Stop_Codons', 'Longest_ORF', 'DNA_Length', 'Is_Best']
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Add all frame data
    row = 2
    for result in all_results:
        is_best = result['Frame'] == best_frames.get(result['Sequence_Name'], '')
        ws_detail.cell(row=row, column=1, value=result['Sequence_Name'])
        ws_detail.cell(row=row, column=2, value=result['Frame'])
        ws_detail.cell(row=row, column=3, value=result['Strand'])
        ws_detail.cell(row=row, column=4, value=result['Protein'])
        ws_detail.cell(row=row, column=5, value=result['Length_AA'])
        ws_detail.cell(row=row, column=6, value=result['Stop_Codons'])
        ws_detail.cell(row=row, column=7, value=result['Longest_ORF'])
        ws_detail.cell(row=row, column=8, value=result['DNA_Length'])
        ws_detail.cell(row=row, column=9, value='Yes' if is_best else '')
        
        # Highlight best frame rows
        if is_best:
            for col in range(1, 10):
                ws_detail.cell(row=row, column=col).fill = PatternFill('solid', fgColor='C6EFCE')
        
        row += 1
    
    # Adjust detail column widths
    ws_detail.column_dimensions['A'].width = 20
    ws_detail.column_dimensions['B'].width = 10
    ws_detail.column_dimensions['C'].width = 10
    ws_detail.column_dimensions['D'].width = 50
    ws_detail.column_dimensions['E'].width = 12
    ws_detail.column_dimensions['F'].width = 12
    ws_detail.column_dimensions['G'].width = 12
    ws_detail.column_dimensions['H'].width = 12
    ws_detail.column_dimensions['I'].width = 10
    
    wb.save(output_file)
    print(f"Results saved to: {output_file}")

def main():
    if len(sys.argv) < 2:
        print("="*60)
        print("DNA to Amino Acid Translator - 6-Frame Analysis")
        print("="*60)
        print("\nUsage: python dna_translator.py <input_file> [output_file]")
        print("\nSupported input formats: .xlsx, .xls, .csv, .tsv")
        print("\nInput file should have columns:")
        print("  - Name/ID column (sequence identifier)")
        print("  - Sequence column (DNA sequence)")
        print("\nExample:")
        print("  python dna_translator.py sequences.xlsx results.xlsx")
        print("  python dna_translator.py sequences.csv")
        print("\nThe tool will:")
        print("  1. Translate each sequence in all 6 reading frames")
        print("  2. Identify the best frame (longest ORF, fewest stops)")
        print("  3. Output results to Excel with summary and detail sheets")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'translation_results.xlsx'
    
    if not os.path.exists(input_file):
        print(f"Error: Input file not found: {input_file}")
        return
    
    print(f"Reading sequences from: {input_file}")
    sequences = read_sequences(input_file)
    print(f"Found {len(sequences)} sequences")
    
    all_results = []
    best_frames = {}
    
    for _, row in sequences.iterrows():
        name = row['Name']
        seq = row['Sequence']
        
        results = analyze_sequence(name, seq)
        if results:
            all_results.extend(results)
            best_frame = find_best_frame(results)
            best_frames[name] = best_frame
            print(f"  {name}: Best frame = {best_frame}")
    
    create_output(all_results, best_frames, output_file)
    print(f"\nProcessed {len(best_frames)} sequences successfully!")

if __name__ == "__main__":
    main()
